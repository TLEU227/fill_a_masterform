#!/usr/bin/env python3
"""Prototyp für Schnittstellen-Option A (siehe
docs/schnittstelle_sysbew_extraktor_vorschlag.md): erzeugt aus
Systembewertungen_GESAMT.xlsx (Blatt 'SysBew') einen Snapshot mit
KLARTEXT-Werten pro System statt der 62 rohen Seriendruck-Checkbox-Spalten.

Anders als db/import_excel.py schreibt dieses Skript NICHT in unsere eigene
DB, sondern erzeugt ein neutrales, von uns unabhängiges Export-Format
(JSON/CSV) - das ist genau das, was wir dem anderen Projekt (SysBew_Extraktor)
als möglichen Vertrag für einen eigenen, deterministischen Export
vorschlagen. Dieses Skript hier ist nur ein Beleg/Prototyp, dass ein solcher
Export aus dem vorhandenen Excel ableitbar ist - keine Aussage darüber, wie
das andere Projekt es letztlich technisch umsetzt.

Deckt alle 16 in docs/analyse_systembewertung_xlsx.md identifizierten
Checkbox-Gruppen ab (nicht nur die Teilmenge, die aktuell in
db/import_excel.py in unsere eigene DB übernommen wird). Für Gruppen mit
bekannten Anomalien (mehrere/keine Markierung, wo eigentlich genau eine
erwartet wird) wird NICHT geraten/automatisch aufgelöst, sondern die
Auffälligkeit explizit im Feld '_anomalien' der Zeile vermerkt (gleiche
Grundhaltung wie im Rest des Projekts: keine erfundenen Aussagen, Lücken
sichtbar machen statt verschweigen).

Nutzung:
    python db/export_normalized_snapshot.py Systembewertungen_GESAMT.xlsx snapshot.json
    python db/export_normalized_snapshot.py Systembewertungen_GESAMT.xlsx snapshot.csv --format csv
    python db/export_normalized_snapshot.py Systembewertungen_GESAMT.xlsx snapshot.json --limit 20
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

import openpyxl

from import_excel import col_index_map, is_marked, get, PLACEHOLDER_PURE_X, PLACEHOLDER_DOKNR

# 1:1-Felder (Freitext/Stammdaten) - unverändert aus db/import_excel.py.DIRECT_MAP,
# hier separat gehalten, damit dieses Skript unabhängig lesbar bleibt.
DIRECT_FIELDS = {
    "MLCSID": "mlcs_id",
    "Betrieb": "bereich",
    "Gebaeude": "gebaeude",
    "Version": "dok_version",
    "Dok. -Nr.": "dok_nummer",
    "AS/BDIS-Name": "systemname",
    "Anlage": "anlage",
    "Raum": "raum",
    "Kurzbeschreibung": "kurzbeschreibung",
    "SW-Name:": "sw_name",
    "SW-Version / Typ:": "sw_version",
    "SW-Hersteller": "sw_hersteller",
    "Hersteller": "hersteller",
    "Phenix": "lieferantennummer",
}

# Die 16 Checkbox-Gruppen aus docs/analyse_systembewertung_xlsx.md.
# typ="single"  -> es wird genau 1 Markierung erwartet (Radiogruppe).
# typ="unabhaengig" -> jede Spalte ist ein eigenes, unabhängiges Ja/Nein-Flag
#                      (bekannt bei QUAL/VAL: 334 Zeilen mit beiden Flags).
GRUPPEN = [
    {"key": "gxp_relevant", "typ": "single",
     "spalten": {"GxP_Relevan_JA": "ja", "GxP_Relevan_NEIN": "nein"}},
    {"key": "gxp_kritikalitaet", "typ": "single",
     "spalten": {"GxP-C": "Critical", "GxP-M": "Major", "GxP-m2": "Minor", "GxP-NA": "N/A"}},
    {"key": "systemtyp", "typ": "unabhaengig",
     "spalten": {"Systemtyp_CIS": "CIS", "Systemtyp_CE": "CE"}},
    {"key": "subtyp", "typ": "single",
     "spalten": {"Subtyp_PCS": "PCS", "Subtyp_LCE": "LCE", "Subtyp_EE": "EE", "Subtyp_NA": "N/A"}},
    {"key": "vnap_stufe", "typ": "single",
     "spalten": {"VNAP_S0": "S0", "VNAP_S1": "S1", "VNAP_S2": "S2"}},
    {"key": "doku_status", "typ": "single",
     "spalten": {"Offen": "offen", "Geschlossen": "geschlossen", "NA": "N/A"}},
    {"key": "gamp_kategorie", "typ": "single",
     "spalten": {"KAT1": "1", "KAT3": "3", "KAT4": "4", "KAT5": "5", "KATNA": "N/A"}},
    {"key": "eres_typ", "typ": "single",
     "spalten": {"ERESTYP1": "Typ 1", "ERESTYP2": "Typ 2", "ERESTYP3": "Typ 3", "ERESTYP4": "Typ 4", "ERESTYPNA": "N/A"}},
    {"key": "testtiefe", "typ": "single",
     "spalten": {"TTIEFEHOCH": "hoch", "TTIEFEMITTEL": "mittel", "TTIEFENIEDRIG": "niedrig"}},
    {"key": "zone_stufe", "typ": "single",
     "spalten": {f"Z{z}S{s}": f"Z{z}S{s}" for z in (1, 2, 3) for s in (1, 2, 3)}},
    {"key": "business_critical", "typ": "single",
     "spalten": {"BCkritisch": "ja", "BCunkritisch": "nein"}},
    {"key": "dokumentart", "typ": "single",
     "spalten": {"Neuerstellung": "Neuerstellung", "Revisioniert": "Revisioniert"}},
    {"key": "geraetekategorie", "typ": "single",
     # Bekanntes Muster (189 Zeilen, siehe docs/analyse_systembewertung_xlsx.md):
     # Subkategorie UND ihre Sammelkategorie werden zusammen markiert (z.B.
     # GKATB2 + GKATB) - das ist KEINE Anomalie, sondern die erwartete
     # Bestaetigung; 'unterkategorie_eltern' macht das explizit, damit nur
     # tatsaechlich widerspruechliche Kombinationen (z.B. zwei Subkategorien,
     # oder Subkategorie + falsche Sammelkategorie) als Anomalie auffallen.
     "spalten": {
         "GKATB1": "B1", "GKATB2": "B2", "GKATB3": "B3",
         "GKATC1": "C1", "GKATC2": "C2",
         "GKATA": "A", "GKATB": "B", "GKATC": "C", "GKATNA": "N/A",
     },
     "unterkategorie_eltern": {
         "GKATB1": "GKATB", "GKATB2": "GKATB", "GKATB3": "GKATB",
         "GKATC1": "GKATC", "GKATC2": "GKATC",
     }},
    {"key": "vq_nvq", "typ": "single",
     "spalten": {"VQ": "VQ", "NVQ": "NVQ"}},
    {"key": "qualifizierung_erforderlich", "typ": "unabhaengig",
     "spalten": {"QUAL": "ja"}},
    {"key": "validierung_erforderlich", "typ": "unabhaengig",
     "spalten": {"VAL": "ja"}},
    {"key": "ki_einstufung", "typ": "single",
     "spalten": {"KI1": "I", "KI2": "II", "KI3": "III", "KI4": "IV", "KI5": "V", "KI6": "VI", "KINA": "N/A"}},
]


def is_placeholder(value: str) -> bool:
    v = value.strip()
    return bool(PLACEHOLDER_PURE_X.match(v) or PLACEHOLDER_DOKNR.match(v))


def markierte_spalten(row: tuple, col_idx: dict, spalten: dict) -> list[str]:
    """Alle Spaltennamen der Gruppe, die im Excel mit 'r' markiert sind
    (roh, ungefiltert - die Basis für Klartext-Wert UND Anomalie-Erkennung)."""
    return [colname for colname in spalten if colname in col_idx and is_marked(row, col_idx, colname)]


def decode_gruppe(row: tuple, col_idx: dict, gruppe: dict, anomalien: list[str], zeile_label: str) -> dict:
    key, typ, spalten = gruppe["key"], gruppe["typ"], gruppe["spalten"]
    markiert = markierte_spalten(row, col_idx, spalten)

    if typ == "unabhaengig" and len(spalten) == 1:
        # z.B. QUAL/VAL: einzelnes Flag, kein Auswahlfeld -> ja/nein, keine Anomalie moeglich.
        (colname,) = spalten.keys()
        return {key: True if markiert else (False if colname in col_idx else None)}

    if typ == "unabhaengig":
        # mehrere unabhaengige Flags in einer Gruppe (z.B. Systemtyp: CIS und CE
        # koennen beide gelten) -> als einzelne Booleans ausgeben, keine Anomalie.
        out = {}
        for colname, wert in spalten.items():
            if colname in col_idx:
                out[f"{key}_{wert.lower()}"] = colname in markiert
        return out

    # typ == "single": genau 1 Markierung erwartet.
    if len(markiert) == 0:
        return {key: None}
    if len(markiert) == 1:
        return {key: spalten[markiert[0]]}

    eltern_map = gruppe.get("unterkategorie_eltern")
    if eltern_map:
        subkats = [c for c in markiert if c in eltern_map]
        andere = [c for c in markiert if c not in eltern_map]
        if len(subkats) == 1 and len(andere) == 1 and eltern_map[subkats[0]] == andere[0]:
            # erwartete Kombination Subkategorie + eigene Sammelkategorie, keine Anomalie
            return {key: spalten[subkats[0]]}

    # Anomalie: mehr als eine Markierung in einer erwarteten Radiogruppe.
    werte = [spalten[c] for c in markiert]
    anomalien.append(f"{zeile_label}: Gruppe '{key}' hat {len(markiert)} Markierungen ({', '.join(werte)}) statt 1")
    return {key: None, f"{key}_mehrfach_markiert": werte}


def export_rows(xlsx_path: Path, limit: int | None = None) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["SysBew"]
    rows = list(ws.iter_rows(values_only=True))
    col_idx = col_index_map(rows[0])
    data_rows = rows[1:]

    ergebnis = []
    stats = {"zeilen": 0, "anomalien": [], "platzhalter_uebersprungen": 0}

    for excel_zeile, row in enumerate(data_rows, start=2):
        systemname = get(row, col_idx, "AS/BDIS-Name")
        mlcs_id = get(row, col_idx, "MLCSID")
        if not systemname and not mlcs_id:
            continue
        if limit is not None and stats["zeilen"] >= limit:
            break

        zeile_label = f"Excel-Zeile {excel_zeile}"
        eintrag = {"excel_zeile": excel_zeile}

        for excel_col, field_key in DIRECT_FIELDS.items():
            wert = get(row, col_idx, excel_col)
            if wert is not None and is_placeholder(wert):
                stats["platzhalter_uebersprungen"] += 1
                wert = None
            eintrag[field_key] = wert

        for gruppe in GRUPPEN:
            eintrag.update(decode_gruppe(row, col_idx, gruppe, stats["anomalien"], zeile_label))

        ergebnis.append(eintrag)
        stats["zeilen"] += 1

    return ergebnis, stats


def write_json(rows: list[dict], path: Path) -> None:
    path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def alle_spaltennamen(rows: list[dict]) -> list[str]:
    # Vereinigung aller Keys, erste Zeile bestimmt die Basis-Reihenfolge,
    # danach angehaengte Zeilen koennen zusaetzliche (Anomalie-)Spalten mitbringen.
    fieldnames = []
    seen = set()
    for r in rows:
        for k in r:
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
    return fieldnames


def write_csv(rows: list[dict], path: Path) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = alle_spaltennamen(rows)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def write_xlsx(rows: list[dict], path: Path) -> None:
    """Bereinigte Master-DB als echtes Excel - gleiche Klartext-Felder wie
    JSON/CSV, aber im Format, das im restlichen Ablauf (Datei-Export statt
    Live-API, siehe docs/architektur_zusammenfuehrung_projekte.md) ohnehin
    ueberall verwendet wird."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterDB_bereinigt"
    if not rows:
        wb.save(path)
        return
    fieldnames = alle_spaltennamen(rows)
    ws.append(fieldnames)
    for r in rows:
        # Listen (z.B. '..._mehrfach_markiert') als lesbaren Text statt Python-Repr
        ws.append([", ".join(r[k]) if isinstance(r.get(k), list) else r.get(k) for k in fieldnames])
    for col in ws.columns:
        letter = col[0].column_letter
        laenge = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(laenge + 2, 60)
    ws.freeze_panes = "A2"
    wb.save(path)


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx_path", help="Pfad zu Systembewertungen_GESAMT.xlsx")
    parser.add_argument("out_path", help="Pfad für die Ausgabedatei (.json, .csv oder .xlsx)")
    parser.add_argument("--format", choices=["json", "csv", "xlsx"], default=None,
                         help="Standard: aus Dateiendung von out_path abgeleitet")
    parser.add_argument("--limit", type=int, default=None, help="nur die ersten N Systeme (zum Testen)")
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx_path)
    out_path = Path(args.out_path)
    suffix = out_path.suffix.lower()
    fmt = args.format or {"csv": "csv", "xlsx": "xlsx"}.get(suffix.lstrip("."), "json")

    rows, stats = export_rows(xlsx_path, limit=args.limit)

    if fmt == "csv":
        write_csv(rows, out_path)
    elif fmt == "xlsx":
        write_xlsx(rows, out_path)
    else:
        write_json(rows, out_path)

    print(f"{stats['zeilen']} Systeme exportiert -> {out_path} ({fmt})")
    print(f"Platzhalterwerte übersprungen: {stats['platzhalter_uebersprungen']}")
    if stats["anomalien"]:
        n = len(stats["anomalien"])
        print(f"\n{n} Anomalien (Mehrfachmarkierung in Radiogruppen), nicht automatisch aufgelöst:")
        for a in stats["anomalien"][:15]:
            print(" -", a)
        if n > 15:
            print(f"   ... und {n - 15} weitere")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
